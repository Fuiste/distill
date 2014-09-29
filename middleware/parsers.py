# -*- coding: utf-8 -*-
import json
import re
from xml.etree import ElementTree
from openpyxl import Workbook, load_workbook

__author__ = "MDee"


def parse_fcc_comments(path):
    """
    """
    root = ElementTree.parse(path)
    text_nodes = root.findall(".//*[@name='text']")
    texts = []
    for t in text_nodes:
        nodes = t[0].text.split("\n")
        t_concat = ""
        for i in range(1, len(nodes)-3):
            t_concat += nodes[i]
        if len(t_concat) < 250:
            texts.append(t_concat)
    texts = sorted(texts, key=lambda t: len(t), reverse=True)
    # texts = filter(lambda t: len(t) < 1000, texts)
    # The length of the one in the middle is 278 chars long
    # Return the top 1000
    return texts[:1000]


def write_xlsx_file(name, contents_list, prompt_text, title):
    wb = Workbook()
    ws = wb.active
    ws.cell("A1").value = prompt_text
    for i in range(0, len(contents_list)):
        ws.cell("A{0}".format(i+2)).value = contents_list[i]
    ws = wb.create_sheet()
    ws.title = title
    wb.save(name)

# texts = parse_fcc_comments(path="/Users/MDee/Desktop/FCC Comments.xml")
# write_xlsx_file(name="/Users/MDee/Desktop/FCC Comments.xlsx", contents_list=texts)


def parse_tweet_json(name):
    json_data=open(name)
    data = json.load(json_data)
    # Strip URLs
    url_regex = r"http(s)?://\w+\.\w+/[\w\d\/_]+"
    # Strip the RT @<name>
    retweet_regex = r"RT\s@[\w\d_]+:"
    statuses = []
    for s in data["statuses"]:
        s2 = re.sub(url_regex, r"", s["text"])
        s2 = re.sub(retweet_regex, r"", s2)
        statuses.append(s2)
    return statuses

def parse_hella_tweets(base_name, count):
    all_tweets = []
    for i in range(1, count+1):
        full_name = base_name.format(i)
        tweets = parse_tweet_json(full_name)
        all_tweets.extend(tweets)
    return all_tweets

# list_of_tweets = parse_hella_tweets(base_name="/Users/MDee/Desktop/Tweet data/ogz{0}.json", count=7)
# write_xlsx_file(name="/Users/MDee/Desktop/OliveGarden tweets.xlsx", contents_list=list_of_tweets, prompt_text="Tweets", title="OliveGarden tweets")

def read_excel_file(name, interesting_column="A"):

    def useless_review(r):
        """
        """
        useless_reviews = ["A new Facebook review was posted to your Page", "A new google review was added", "This review is hidden. View review"]
        for u in useless_reviews:
            if r == u:
                return True
        return False

    book = load_workbook(name)
    sheet = book[book.get_sheet_names()[0]]
    vals= []
    for i in range(2, len(sheet.columns[0])):
        v = sheet["{0}{1}".format(interesting_column, i)].value
        if len(v) and not useless_review(v.encode("utf8")):
            vals.append(v)
    return vals

def write_json_file(name, content):
    with open(name, "w") as outfile:
        json.dump(content, outfile)

vals = read_excel_file(name="/Users/MDee/Desktop/Yum-Brands-Sample-Reviews-Report-09-15-2014.xlsx", interesting_column="J")
write_json_file(name="/Users/MDee/Desktop/yum-json.json", content=vals)